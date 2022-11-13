from flask import Flask, jsonify, request, abort, render_template, redirect, url_for, send_file
import os.path
from openpyxl import Workbook, load_workbook
import datetime
import requests
from db import connection
import json


app = Flask(__name__)

 
@app.route('/billing-api/health')
def health_db_status():
    try:
        
        with connection.cursor() as mycursor:
            mycursor = connection.cursor(dictionary=True)
            stmt = "select 1"
            mycursor.execute(stmt)
            stmt_result = mycursor.fetchone()
            return jsonify({"status": "OK"}), 200

    except Exception as e:
        
        return jsonify({"status":"failure"}), 500
    


@app.route('/provider', methods=['GET', 'POST'])
def provider():
    if request.method == 'POST':
        body = request.get_json()
        name1 = body['name']
        if name1 != None:
            with connection.cursor() as mycursor:
                mycursor = connection.cursor(dictionary=True)
                stmt,val = "INSERT INTO Provider (name) VALUES (%s)",[(name1)]
                print (name1)
                print(stmt)
                mycursor.execute(stmt, val)

                connection.commit()
                return jsonify({"record saves provider id: ": mycursor.lastrowid}), 201
        else:
            return jsonify({"msg": " Unsuccessfull!!!"}), 204
            
    else:

        with connection.cursor() as mycursor:
            mycursor = connection.cursor(dictionary=True)
            do = "SELECT * FROM Provider"
            mycursor.execute(do)
            result = mycursor.fetchall()

            return jsonify(result)


@app.route('/provider/<id>', methods=['GET', 'PUT'])
def update_provider_name(id):
    if request.method == 'PUT':

        body = request.get_json()
        name = body['name']

        if name != '':
            with connection.cursor() as mycursor:
                mycursor= connection.cursor(dictionary=True)
                do,val = "UPDATE Provider SET name =%s where id=%s",(name,id)
                mycursor.execute(do,val)
                mycursor.commit()
                return jsonify({"message":"update succes:  "}), 201
        else:
            return jsonify({"msg": " Unsuccessfull!!!"}), 204


    if request.method == 'GET':
             with connection.cursor() as mycursor:
                 do = "SELECT * FROM Provider where id=%s"
                 mycursor = connection.cursor(dictionary=True)
                 mycursor.execute(do,[(id)])
                 result = mycursor.fetchone()

                 return jsonify(result)
            # retrieve specific provider here 

             pass



@app.route('/weight', methods = ["POST"])
def add_weight():



    body = request.get_json()

    data = {
        "container": body['container'],
        "weight": body['weight'],
        "truck_id": body['truck_id'],
        "unit": body['unit'],
        "force": body['force']
    }


    return jsonify(data), 201


@app.route('/rates', methods=['GET', 'POST'])
def rates():
    if request.method == 'POST':
        # receive post parameters
        body = request.get_json()
        # name = body['name']
        file = body['file']
        filepath = f'./in/{file}'
        # check if file exist
        if os.path.isfile(filepath):
            # df = pd.read_excel(file)
            data = []
            df = load_workbook(filepath)
            sheet = df.active
            row_count = sheet.max_row
            for rows in sheet.iter_rows():
                row_cells = []
                for cell in rows:
                    row_cells.append(cell.value)
                data.append(tuple(row_cells))
                
            # establish db connection
            with connection.cursor() as mycursor:
                # mycursor = connection.cursor(dictionary=True)
                for row in data[1:]:
                    mycursor.execute("SELECT * FROM Rates WHERE product_id = %s AND scope = %s", (row[0], row[2], ))
                    result = mycursor.fetchall()
                    if len(result) < 1:
                        mycursor.execute("INSERT INTO Rates (`product_id`, `rate`, `scope`) VALUES (%s, %s, %s)", row)
                    else:
                        mycursor.execute("""
                        UPDATE Rates 
                        SET rate = %s 
                        WHERE product_id = %s AND scope = %s
                        """, (row[1], f'{row[0]}', f'{row[2]}'))
                    connection.commit()
                return jsonify({"msg": "Rates uploaded successfully!"}), 201
        else:
            return jsonify({"msg": "Unsupported file format!!!"}), 204
    else:

        if "display" not in request.args:

            with connection.cursor() as mycursor:
                #mycursor = connection.cursor(dictionary=True)
                stmt = "SELECT * FROM Rates"
                mycursor.execute(stmt)
                stmt_result = mycursor.fetchall()
                # export data to excel file
                wb = Workbook()
                ws = wb.active
                ws.title = "rates"

                ws.append(['Product', 'Rate', 'Scope'])

                for row in stmt_result:
                    ws.append(list(row))
                wb.save(f"in/rates-{datetime.datetime.now().strftime('%Y%m%d')}.xlsx")
                
                return jsonify({"msg": "Rates downloaded successfully!"})

        else:

            # Get all rates to frontend
            with connection.cursor() as mycursor:
                mycursor = connection.cursor(dictionary=True)
                do = "SELECT * FROM Rates"
                mycursor.execute(do)
                result = mycursor.fetchall()
                return jsonify(result)


@app.route('/bill/<id>')
def getbill(id):
    t1 = request.args.get('t1') if request.args.get('t1') else datetime.datetime(2022,1,1).strftime('%Y%m%d%H%M%S')
    t2 = request.args.get('t2') if request.args.get('t2') else datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    param={"from": t1,"to": t2}

    sessionuri = requests.get(f"http://ec2-18-192-110-37.eu-central-1.compute.amazonaws.com:8081/session/{id}").json()
    # weighturi = requests.get(f"http://ec2-18-192-110-37.eu-central-1.compute.amazonaws.com:8081/weight/{id}", params=param).json()
    
    # expected return
    return jsonify({
        "id": sessionuri[0]['id'],
        "name": "<str>",
        "from": sessionuri[0]['truck'],
        "to": sessionuri[0]['truck'],
        "truckCount": sessionuri[0]['truck'],
        "sessionCount": sessionuri[0]['truck'],
        "products": [{ "product":"<str>","count": "<str>", "amount": "<int>", "rate": "<int>", "pay": "<int>"}],
        "total": sessionuri[0]['neto'] 
    })



#Endpoint for post truck    
@app.route('/truck', methods=["GET",'POST'])
def Truck_Post():
    if request.method == 'POST':
        body = request.get_json()
        truck_id = body['id']
        provider_id = body['provider_id']

        if truck_id != '' and provider_id != None :
                try: 
                    with connection.cursor() as mycursor:
                        mycursor = connection.cursor(dictionary=True)
                        stmt, val = "INSERT INTO Trucks (id, provider_id) VALUES (%s, (select id from Provider where id=%s))", (truck_id, provider_id)
                        mycursor.execute(stmt, val)
                        connection.commit()
                        return jsonify({"message": "data saved successfully!"}), 201
                except Exception as e:
                    return jsonify({"message": "failure posting "}), 400
        else:
            return jsonify({"message": "provide Truck ID and Provider ID "}), 400

    elif request.method == "GET":
        # Get all trucks
        with connection.cursor() as mycursor:
            mycursor = connection.cursor(dictionary=True)
            do = "SELECT * FROM Trucks"
            mycursor.execute(do)
            result = mycursor.fetchall()
            return jsonify(result)

    else:
        return jsonify({"message": "method not allowed"}), 405



#End point for put track_id
@app.route('/truck/<id>', methods=['PUT'])
def Truck_Put(id):
    
    if request.method == 'PUT':
        body = request.get_json()
        truck_id = request.args.get('id')
        provider_id = body['provider_id']

        if id !='' and provider_id != None:    
            with connection.cursor() as mycursor:
                mycursor = connection.cursor(dictionary=True)
                stmt = "update Trucks set provider_id = %s where id=%s" 
                val=(provider_id, id)
                mycursor.execute(stmt,val)
                mycursor.commit
                return jsonify ({"message": "provider ID updated successfully! "}), 201
    else:
            return jsonify({"msg": "Truck ID not found in the database "}), 204


@app.route('/truck/<id>', methods=['GET'])
def get_truckid(id):

    t1 = request.args.get('t1')
    t2 = request.args.get('t2')
    if id != None and t1 and t2:
        try:
        
            param={'id':id,"from":t1,"to":t2}
            reqResp=requests.get("http://ec2-18-192-110-37.eu-central-1.compute.amazonaws.com:8081/",params=param)
            assert reqResp.status_code == 200
            data=reqResp.json()
            return data

            
        except Exception as e:
                    return jsonify({"message": "failure fetching data "}), 400
    else:

        # Get truck record for update
        with connection.cursor() as mycursor:
            do = "SELECT * FROM Trucks where id=%s"
            mycursor = connection.cursor(dictionary=True)
            mycursor.execute(do,[(id)])
            result = mycursor.fetchone()

            return jsonify(result)

            
        





### BEGIN FRONTEND ROUTE ###

@app.route('/', methods = ["GET"])
def billing_index():
    try:
        provider_count = 0
        truck_count = 0
        rate_count = 0


        # Get all rates to frontend
        with connection.cursor() as mycursor:
            mycursor = connection.cursor(dictionary=True)

            # Rate
            rate_query = "SELECT * FROM Rates"
            mycursor.execute(rate_query)
            rate_count = len(mycursor.fetchall())

            # Truck
            truck_query = "SELECT * FROM Trucks"
            mycursor.execute(truck_query)
            truck_count = len(mycursor.fetchall())

            # Provider
            provider_query = "SELECT * FROM Provider"
            mycursor.execute(provider_query)
            provider_count = len(mycursor.fetchall())



        return render_template(
            'index.html',
            is_dashboard = True,
            provider_count  = provider_count,
            truck_count  = truck_count,
            rate_count  = rate_count
            )
    except Exception as err:
        abort(500)

@app.route('/provider-list', methods = ["GET"])
def get_provider_list():
    try:

        return render_template('provider-list.html', is_provider=True)
    except Exception as err:
        abort(500)


@app.route('/truck-list', methods = ["GET"])
def get_truck_list():
    try:

        result = []

        with connection.cursor() as mycursor:
            mycursor = connection.cursor(dictionary=True)
            do = "SELECT * FROM Provider"
            mycursor.execute(do)
            result = mycursor.fetchall()

        return render_template('truck-list.html',
        is_truck=True,
        providers = result
        )
    except Exception as err:
        abort(500)


@app.route('/rate-list', methods = ["GET"])
def get_rate_list():
    try:

        return render_template('rate-list.html', is_rate=True)
    except Exception as err:
        abort(500)




@app.route('/download/rates')
def downloadFile ():
    path = f"in/rates-{datetime.datetime.now().strftime('%Y%m%d')}.xlsx"
    with connection.cursor() as mycursor:
        #mycursor = connection.cursor(dictionary=True)
        stmt = "SELECT * FROM Rates"
        mycursor.execute(stmt)
        stmt_result = mycursor.fetchall()
        # export data to excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "rates"

        ws.append(['Product', 'Rate', 'Scope'])

        for row in stmt_result:
            ws.append(list(row))
        wb.save(path)


    return send_file(path, as_attachment=True)

### BEGIN FRONTEND ROUTE ###


@app.errorhandler(500)
def internal_server_error(error):
    return jsonify({
        "success": False,
        "error": 500,
        "message": "Internal Server Error"
    })




if __name__ == '__main__':
    app.run(debug=True, port=5001)